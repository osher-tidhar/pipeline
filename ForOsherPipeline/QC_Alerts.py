import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import gspread
import json
from oauth2client.service_account import ServiceAccountCredentials
import datetime
from glob import glob
import polars as pl
from openpyxl import load_workbook
from collections import deque
import re

base_path = '/home/ubuntu/pipeline/ForOsherPipeline'


def send_email_alert(subject, body, to_email):
    from_email = "osher@respirai.com"
    from_password = "yewy ethk okqh jnmj"  # use app password or environment variable

    # Build the email
    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))

    # Connect to Gmail's SMTP server
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()  # Secure the connection
        server.login(from_email, from_password)
        server.send_message(msg)
        server.quit()
        print("✅ Alert sent successfully.")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")


def send_data_analyzed_qc_alert(patient_ID, date):
    file_path = f"/home/ubuntu/pipeline/outputs/{patient_ID}/final_step_{date}_output.txt"
    
    if not os.path.exists(file_path):
        subject = f"Pipeline QC Alert: Missing Analysis Output for Patient {patient_ID}"
        body = (
            f"<b><span style='color:red;'>Patient ID: {patient_ID}</span></b><br>"
            f"Expected analysis file was not found for date: <b>{date}</b>.<br><br>"
            f"Missing file path:<br><code>{file_path}</code><br><br>"
            f"Please check the processing pipeline or logs for any errors."
            f"<i>This is an automated QC monitoring alert.</i>"
        )

        send_email_alert(subject, body, to_email="osher@respirai.com" \
            ", nadav@respirai.com, vladimir@respirai.com, dror@respirai.com" \
            ""
        )
    else:
        print(f"Pipeline QC Log: ✅ Analysis done for {patient_ID} on {date}")


def read_excel_as_polars(path):
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    data = list(sheet.values)
    headers = data[0]
    rows = data[1:]
    return pl.DataFrame([dict(zip(headers, row)) for row in rows])

def dtw_alerts(path_excel_file):
    try:
        wb = load_workbook(path_excel_file, data_only=True)
    except Exception as e:
        print(f"❌ Failed to read Excel file: {e}")
        return

    noise_history = {}
    ecg_history = {}

    for sheet_name in wb.sheetnames:
        patient_id = sheet_name.strip()
        sheet = wb[sheet_name]

        # Read data from this sheet
        data = list(sheet.values)
        if not data or len(data) < 2:
            continue  # skip if sheet is empty or only headers

        headers = data[0]
        rows = data[1:]
        df = pl.DataFrame([dict(zip(headers, row)) for row in rows])

        noise_history.setdefault(patient_id, deque(maxlen=3))
        ecg_history.setdefault(patient_id, deque(maxlen=4))

        for i, row in enumerate(df.iter_rows(named=True)):
            date = row.get("date")
            if not date:
                continue

            if isinstance(date, datetime.datetime):
                date = date.strftime("%Y-%m-%d")

            triggers = []

            # 1. Condition == "BAD"
            if i == 0:
                condition = str(row.get("Condition", "") or "").strip().upper()
                if condition == "BAD":
                    triggers.append("Patch Position ('Condition') is BAD")
                    
            # 2. Noise > 25% for 3 consecutive days
            if i < 3:
                try:
                    noise_val = float(row.get("Noise", 0) or 0)
                    noise_history[patient_id].append(noise_val > 25)
                except:
                    noise_history[patient_id].append(False)

                # After the third row, evaluate if all past 3 days had >25% noise
                if i == 2 and all(noise_history[patient_id]):
                    triggers.append("Noise > 25% in the past 3 days")
            
            # 3. ECG Percent < 50% for 4 consecutive days
            if i < 4:
                try:
                    ecg_val = row.get("ECG Percent")
                    ecg_val = ecg_val[:-1]
                    ecg_val = float(ecg_val or 100)
                    ecg_history[patient_id].append(ecg_val < 50)
                except:
                    ecg_history[patient_id].append(False)

                if i == 3 and all(ecg_history[patient_id]):
                    triggers.append("ECG Percent < 50% for the past 4 consecutive days")

            # 4. SpO2 Sessions (4-hour apart) == 0
            if i == 0:
                try:
                    spo2_val = int(row.get("SpO2 Sessions (4-hour apart)", 1) or 1)
                    if spo2_val == 0:
                        triggers.append("SpO2 Sessions (4-hour apart) == 0")
                except:
                    pass

            if triggers:
                body = f"""
                <html>
                <body>
                    <p><strong>Data Alert for Patient:</strong> <span style='color:blue; font-weight:bold'>{patient_id}</span></p>
                    <p><strong>Date:</strong> {date}</p>
                    <ul>
                        {''.join(f"<li>{t}</li>" for t in triggers)}
                    </ul>
                </body>
                </html>
                """
                body = body + f"<br><i>This is an automated QC monitoring alert.</i>"

                send_email_alert(subject=f"Pipeline QC Data Alert: Patient {patient_id}", body=body, to_email="osher@respirai.com" \
                ", dror@respirai.com, noam@respirai.com, shai@respirai.com" \
                ""
            )


def send_exacerbation_alert():
    # Setup
    SHEET_KEY = '1fl0AS6iyDHFK1DAx_6o8M6DxFRg9HFm37sjAf-WsUCg'
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name(base_path + '/qc-alerts-456909-116763b16ded.json', scope)
    gc = gspread.authorize(creds)
    sheet = gc.open_by_key(SHEET_KEY).sheet1
    data = sheet.get_all_records()

    # Load previous state
    try:
        with open('last_qc_check.json', 'r') as f:
            previous = json.load(f)
    except FileNotFoundError:
        previous = {}

    alerts = []

    # Process each row
    for row in data:
        patient_id = row.get("Patient ID", "").strip()
        
        # Skip rows with missing or empty patient ID
        if not patient_id or not re.search(r'\d', patient_id):
            continue
        
        sev_count = row.get("# of exacerbations during the study - severe", 0)
        mod_count = row.get("# of exacerbations during the study - moderate", 0)

        # Skip alert if this is a new patient not seen before
        if patient_id not in previous:
            previous[patient_id] = {"severe": sev_count, "moderate": mod_count}
            continue
        
        old_sev = previous.get(patient_id, {}).get("severe", 0)
        old_mod = previous.get(patient_id, {}).get("moderate", 0)

        if sev_count != old_sev or mod_count != old_mod:
            alert = (
                f"<b><span style='color:#1E90FF;'>Patient ID: {patient_id} - </span></b>"
                f"Severe: {old_sev} ➝ {sev_count}, Moderate: {old_mod} ➝ {mod_count}<br>"
            )

            date_sev = row.get("Severe - Dates", "N/A")
            date_mod = row.get("Moderate - Dates", "N/A")
            alert += f"\n   Dates (severe): {date_sev}<br>"
            alert += f"\n   Dates (moderate): {date_mod}<br>"

            alerts.append(alert)

        # Save updated counts
        previous[patient_id] = {"severe": sev_count, "moderate": mod_count}

    # Send email if anything changed
    if alerts:
        body = "".join(alerts)
        body += f"<br><i>This is an automated QC monitoring alert.</i>"
        send_email_alert(
            subject="Pipeline QC Alert: New Exacerbation Detected",
            body=body,
            to_email="osher@respirai.com" \
            #", nadav@respirai.com, vladimir@respirai.com, dror@respirai.com" \
            ""
        )

    # Save state
    with open('last_qc_check.json', 'w') as f:
        json.dump(previous, f)


def check_for_stalled_patients(base_dir, patient_list, days_threshold=3):
    today = datetime.date.today()
    cutoff_date = today - datetime.timedelta(days=days_threshold)

    for patient_id in patient_list:
        # Example file pattern: ECGRec_202405_E117969_2025-03-19_ECG_denoised.csv
        pattern = os.path.join(base_dir, patient_id, "patch_files", "ECGRec_*_ECG_denoised.csv")
        files = glob(pattern)
        
        if not files:
            send_stalled_patient_alert(patient_id, "No files found")
            continue

        # Extract dates from file names
        latest_date = None
        for file in files:
            parts = os.path.basename(file).split("_")
            if len(parts) >= 5:
                try:
                    date_str = parts[3]
                    file_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                    if not latest_date or file_date > latest_date:
                        latest_date = file_date
                except Exception:
                    continue

        if latest_date and latest_date < cutoff_date:
            send_stalled_patient_alert(patient_id, latest_date)


def send_stalled_patient_alert(patient_ID, last_date):
    subject = f"Pipeline QC Alert: 🛑 No Recent Data Alert: Patient {patient_ID}"
    body = (
        f"<b><span style='color:darkred;'>Patient ID: {patient_ID}</span></b><br>"
        f"No new data received since <b>{last_date}</b>.<br>"
        f"Please check for device issues or patient compliance.<br><br>"
        f"<i>This is an automated QC monitoring alert.</i>"
    )
    send_email_alert(subject, body, to_email="osher@respirai.com" \
            ", dror@respirai.com, shai@respirai.com, noam@respirai.com" \
            ""
    )
